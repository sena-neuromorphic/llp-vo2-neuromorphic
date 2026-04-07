import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from pathlib import Path
from scipy.signal import butter, filtfilt
from openpyxl import load_workbook

from pathlib import Path
import pandas as pd

HERE = Path().resolve()  # diretório atual no Colab
XLSX = HERE / "Experimental_Data_VO2.xlsx"

df = pd.read_excel(XLSX)

# ============================================================
# LLP FUNCTIONS
# ============================================================
def ProFuncVO2(x, gama):
    return 0.5 * (1.0 - np.sin(gama * x)) * (1.0 + np.tanh(np.pi**2 - 2.0*np.pi*x))

def FL(Tval, delta, w, Tc, beta):
    return 0.5 + 0.5 * np.tanh(beta * (delta * w / 2.0 + Tc - Tval))

def anchor_major_auto(T0, T1, w, Tc, beta, gama, Tpr0=None, eps_g=1e-6):
    delta = +1 if (T1 - T0) > 0 else -1
    Tr = float(T0)
    if Tpr0 is None:
        Tpr0 = float(w)
    P0 = ProFuncVO2(0.0, gama)
    arg0 = delta * w / 2.0 + Tc - (Tr + Tpr0 * P0)
    gr = float(np.clip(0.5 + 0.5 * np.tanh(beta * arg0), eps_g, 1.0 - eps_g))
    return Tr, gr, float(Tpr0), delta

def update_hysteresis(T, w, Tc, beta, gama, Tpr0=None, eps_dT=1e-9):
    N = len(T)
    g = np.zeros(N); delta_vec = np.zeros(N, dtype=int)
    Tr_vec = np.zeros(N); Tpr_vec = np.zeros(N)

    Tr, gr, Tpr, delta = anchor_major_auto(T[0], T[1], w, Tc, beta, gama, Tpr0=Tpr0)
    g[0] = gr; delta_vec[0] = delta; Tr_vec[0] = Tr; Tpr_vec[0] = Tpr

    for n in range(1, N):
        dT = T[n] - T[n-1]
        delta_new = +1 if dT > eps_dT else (-1 if dT < -eps_dT else delta)
        if delta_new != delta:
            Tr = T[n-1]
            gr = float(np.clip(g[n-1], 1e-6, 1.0 - 1e-6))
            delta = delta_new
            P0 = ProFuncVO2(0.0, gama)
            Tpr = (delta * w / 2.0 + Tc - Tr - np.arctanh(2.0 * gr - 1.0) / beta) / max(P0, 1e-9)
            if abs(Tpr) < 1e-9:
                Tpr = np.sign(Tpr) * 1e-9 if Tpr != 0 else 1e-3
        Tpr_safe = np.sign(Tpr) * max(abs(Tpr), 1e-9)
        x = (T[n] - Tr) / Tpr_safe
        arg = delta * w / 2.0 + Tc - (T[n] + Tpr_safe * ProFuncVO2(x, gama))
        g[n] = float(np.clip(0.5 + 0.5 * np.tanh(beta * arg), 0.0, 1.0))
        delta_vec[n] = delta; Tr_vec[n] = Tr; Tpr_vec[n] = Tpr

    return g, delta_vec, Tr_vec, Tpr_vec

# ============================================================
# LLP PARAMETERS
# ============================================================
w     = 6.67
Tc    = 47.9
beta  = 0.212
gama  = 0.9
R0    = 17
Rm    = 140.0

# ============================================================
# EXCEL READING
# ============================================================
print("Readin data from Excel...")
df_all = pd.read_excel(
    XLSX,
    sheet_name='FORCs',
    header=2,
    skiprows=[3],
    usecols=['T_ref (°C)', 'T_exp (°C)', 'R_exp (Ω)', 'Índice segmento']
)
print(f"Total samples loaded: {len(df_all)}")

T_ref_all = df_all['T_ref (°C)'].to_numpy(dtype=float)
T_raw_all = df_all['T_exp (°C)'].to_numpy(dtype=float)
R_raw_all = df_all['R_exp (Ω)'].to_numpy(dtype=float)

# ============================================================
# BUTTERWORTH FILTER ON ALL DATA (FORCs included)
# ============================================================
FS = 1.25          # Hz  (Ts = 0.8 s)
FC = 0.15          # Hz
Wn = FC / (FS / 2.0)

b, a = butter(2, Wn, btype='low', analog=False)
T_filt_all = filtfilt(b, a, T_raw_all)
R_filt_all = filtfilt(b, a, R_raw_all)

print(f"Butterworth 2ª order, fc={FC} Hz | Wn={Wn:.4f}")
print(f"T_exp Filter (all): [{T_filt_all.min():.3f}, {T_filt_all.max():.3f}] °C")
print(f"R_exp Filter (all): [{R_filt_all.min():.2f}, {R_filt_all.max():.2f}] Ω")

# ============================================================
# LLP ON ALL DATA (FORCs included)
# ============================================================
print("Running LLP on all data (FORCs)...")
g_all, _, _, _ = update_hysteresis(T_ref_all, w, Tc, beta, gama, Tpr0=w)
R_model_all = np.maximum(g_all * R0 * np.exp(2553.0 / (T_ref_all + 273.0)) + Rm, 1e-9)
print(f"R_model LLP (all): [{R_model_all.min():.2f}, {R_model_all.max():.2f}] Ω")

# ============================================================
# MAJOR LOOPS ONLY (segments 0 and 1) — for Excel + TCR
# ============================================================
mask_ml = df_all['Índice segmento'].isin([0, 1])
df_ml   = df_all[mask_ml].reset_index(drop=True)
idx_ml  = df_all.index[mask_ml].to_numpy()   # original indices for filter slicing

T_ref  = df_ml['T_ref (°C)'].to_numpy(dtype=float)
T_raw  = df_ml['T_exp (°C)'].to_numpy(dtype=float)
R_raw  = df_ml['R_exp (Ω)'].to_numpy(dtype=float)
T_filt = T_filt_all[idx_ml]
R_filt = R_filt_all[idx_ml]
print(f"Major loops only: {len(df_ml)} samples  "
      f"(T_ref [{T_ref.min():.1f}, {T_ref.max():.1f}] °C)")

# LLP on major loops
print("Running LLP on major loops...")
g, _, _, _ = update_hysteresis(T_ref, w, Tc, beta, gama, Tpr0=w)
R_model = np.maximum(g * R0 * np.exp(2553.0 / (T_ref + 273.0)) + Rm, 1e-9)
print(f"R_model LLP (major loops): [{R_model.min():.2f}, {R_model.max():.2f}] Ω")

# ============================================================
# Saving EXCEL
# ============================================================
# Mapeamento de colunas (índice 1-based):
#   G=7  → T_expFilter (°C)
#   J=10 → R_expFilter (Ω)
#   K=11 → R_model (Ω)
COL_TFILT  = 7
COL_RFILT  = 10
COL_RMOD   = 11
ROW_START  = 5

print("Salving T_expFilter R_expFilter R_model Excel...")
OUT_XLSX = XLSX

wb = load_workbook(OUT_XLSX)
ws = wb['FORCs']
for i in range(len(T_ref)):
    row = ROW_START + i
    ws.cell(row=row, column=COL_TFILT, value=round(float(T_filt[i]),   9))
    ws.cell(row=row, column=COL_RFILT, value=round(float(R_filt[i]),   6))
    ws.cell(row=row, column=COL_RMOD,  value=round(float(R_model[i]), 6))
wb.save(OUT_XLSX)
print(f"File saved: {OUT_XLSX}")

# ============================================================
# PLOT SETTINGS
# ============================================================
plt.rcParams.update({
    'font.family':      'serif',
    'font.size':        9,
    'axes.labelsize':   9,
    'xtick.labelsize':  8,
    'ytick.labelsize':  8,
    'legend.fontsize':  7.5,
    'lines.linewidth':  0.9,
    'axes.linewidth':   0.8,
    'xtick.direction':  'in',
    'ytick.direction':  'in',
    'xtick.top':        True,
    'ytick.right':      True,
    'grid.linewidth':   0.5,
    'grid.alpha':       0.4,
})

# ============================================================
# PLOT 1 — Raw experimental vs LLP model
# ============================================================
FIG_W, FIG_H = 7.0, 5.6

fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))

ax.semilogy(T_raw_all, R_raw_all,
            color='#1f77b4', linewidth=0.9, rasterized=True,
            label=r'Experimental Raw')

ax.semilogy(T_ref_all, R_model_all,
            color='#ff7f0e', linewidth=0.9, rasterized=True,
            label=r'LLP model')

ax.set_xlabel(r'$T$ (°C)')
ax.set_ylabel(r'$R$ ($\Omega$)')
ax.set_xlim(19, 81)
ax.xaxis.set_major_locator(ticker.MultipleLocator(10))
ax.xaxis.set_minor_locator(ticker.MultipleLocator(5))
ax.yaxis.set_major_locator(ticker.LogLocator(base=10, numticks=6))
ax.yaxis.set_minor_locator(
    ticker.LogLocator(base=10, subs=np.arange(2, 10) * 0.1, numticks=60))
ax.yaxis.set_major_formatter(ticker.LogFormatterMathtext())
ax.grid(True, which='major', linestyle='--')
ax.grid(True, which='minor', linestyle=':', linewidth=0.3, alpha=0.3)
ax.legend(loc='upper right', framealpha=0.9)

fig.tight_layout(pad=0.4)
fig.savefig(HERE / 'Experimental_Model.pdf', dpi=300, bbox_inches='tight')
fig.savefig(HERE / 'Experimental_Model.png', dpi=300, bbox_inches='tight')
print("Saved figures:")
print("  Experimental_Model.pdf")
print("  Experimental_Model.png")
plt.show()

# ============================================================
# PLOT 2 — Filtered experimental vs LLP model
# ============================================================
FIG_W, FIG_H = 7.0, 5.6

fig, ax = plt.subplots(figsize=(FIG_W, FIG_H))

ax.semilogy(T_filt_all, R_filt_all,
            color='#1f77b4', linewidth=0.9, rasterized=True,
            label=r'Experimental Filtered')

ax.semilogy(T_ref_all, R_model_all,
            color='#ff7f0e', linewidth=0.9, rasterized=True,
            label=r'LLP model')

ax.set_xlabel(r'$T$ (°C)')
ax.set_ylabel(r'$R$ ($\Omega$)')
ax.set_xlim(19, 81)
ax.xaxis.set_major_locator(ticker.MultipleLocator(10))
ax.xaxis.set_minor_locator(ticker.MultipleLocator(5))
ax.yaxis.set_major_locator(ticker.LogLocator(base=10, numticks=6))
ax.yaxis.set_minor_locator(
    ticker.LogLocator(base=10, subs=np.arange(2, 10) * 0.1, numticks=60))
ax.yaxis.set_major_formatter(ticker.LogFormatterMathtext())
ax.grid(True, which='major', linestyle='--')
ax.grid(True, which='minor', linestyle=':', linewidth=0.3, alpha=0.3)
ax.legend(loc='upper right', framealpha=0.9)

fig.tight_layout(pad=0.4)
fig.savefig(HERE / 'Experimental_Filtered_Model.pdf', dpi=300, bbox_inches='tight')
fig.savefig(HERE / 'Experimental_Filtered_Model.png', dpi=300, bbox_inches='tight')
print("Saved figures:")
print("  Experimental_Filtered_Model.pdf")
print("  Experimental_Filtered_Model.png")
plt.show()

# ============================================================
# TCR CALCULATION
# TCR = (1/R) * dR/dT  [°C⁻¹]
# Central difference over a 100-sample window (HALF = 50):
#   dR/dT|n ≈ (R[n+50] - R[n-50]) / (T[n+50] - T[n-50])
# First and last 50 samples are left as NaN.
# Only negative values in (-0.5, 0) are retained.
# ============================================================
HALF = 50   # half-window → total window of 100 samples

def grad100(T_arr, R_arr):
    N   = len(T_arr)
    tcr = np.full(N, np.nan)
    for n in range(HALF, N - HALF):
        dT_left  = T_arr[n]        - T_arr[n - HALF]
        dT_right = T_arr[n + HALF] - T_arr[n]
        # Reject window if the two halves move in opposite directions (reversal inside)
        if dT_left * dT_right <= 0:
            continue
        dT = T_arr[n + HALF] - T_arr[n - HALF]
        dR = R_arr[n + HALF] - R_arr[n - HALF]
        if abs(dT) > 1e-6:
            tcr[n] = (dR / dT) / R_arr[n]
    return tcr

TCR_exp_raw   = grad100(T_filt,  R_filt)
TCR_model_raw = grad100(T_ref,   R_model)

def clip_neg(tcr):
    result = tcr.copy()
    result[(result >= 0) | (result < -0.5)] = np.nan
    return result

TCR_exp   = clip_neg(TCR_exp_raw)
TCR_model = clip_neg(TCR_model_raw)

# ============================================================
# PLOT 3 — TCR: Experimental Filtered vs LLP model
# ============================================================
fig3, ax3 = plt.subplots(figsize=(7.0, 5.6))

ax3.plot(T_filt, TCR_exp,
         color='#1f77b4', linewidth=0.9, rasterized=True,
         label=r'TCR Experimental Filtered')

ax3.plot(T_ref, TCR_model,
         color='#ff7f0e', linewidth=0.9, rasterized=True,
         label=r'TCR LLP model')

ax3.axhline(0, color='black', linewidth=0.6, linestyle='--')
ax3.set_ylim(-0.55, 0.05)

ax3.set_xlabel(r'$T$ (°C)')
ax3.set_ylabel(r'TCR ($^{\circ}\mathrm{C}^{-1}$)')
ax3.set_xlim(19, 81)
ax3.xaxis.set_major_locator(ticker.MultipleLocator(10))
ax3.xaxis.set_minor_locator(ticker.MultipleLocator(5))
ax3.yaxis.set_major_locator(ticker.MaxNLocator(6))
ax3.yaxis.set_minor_locator(ticker.AutoMinorLocator(5))
ax3.grid(True, which='major', linestyle='--')
ax3.grid(True, which='minor', linestyle=':', linewidth=0.3, alpha=0.3)
ax3.legend(loc='lower right', framealpha=0.9)

fig3.tight_layout(pad=0.4)
fig3.savefig(HERE / 'TCR_Experimental_Filtered_Model.pdf', dpi=300, bbox_inches='tight')
fig3.savefig(HERE / 'TCR_Experimental_Filtered_Model.png', dpi=300, bbox_inches='tight')
print("Saved figures:")
print("  TCR_Experimental_Filtered_Model.pdf")
print("  TCR_Experimental_Filtered_Model.png")
plt.show()
